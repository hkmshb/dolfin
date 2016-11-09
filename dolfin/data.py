# -*- coding: utf-8 -*-
"""
Defines functions and classes for data access operations.
"""
import os
from abc import ABCMeta, abstractmethod
from .core import Storage



class Db:

    @staticmethod
    def make_dml_provider(dml_builder, row_provider):
        def dml_provider():
            for row in row_provider:
                try:
                    dml = dml_builder(row)
                    yield dml
                except Exception as ex:
                    raise ex
        return dml_provider
    
    @staticmethod
    def make_row_provider(conn, table_name, columns=None, 
        extra_clause=None, count=None):
        # build query text
        query = "SELECT %s FROM %s" % (
            '*' if not columns else ', '.join(columns),
            table_name
        )

        if extra_clause:
            query += extra_clause
        
        def read_rows():
            # execute query
            cursor = conn.cursor()
            cursor.execute()

            desc = cursor.description
            fields = [f[0] for f in desc]

            records = cursor.fetchmany(count) if count else cursor.fetchall()
            for r in records:
                yield _(zip(fields, r))
        return read_rows()

    @staticmethod
    def process(self, conn, dml_provider, commit_interval=10,
        on_dml_processed=None):
        if not dml_provider:
            raise ValueError("dml_provider must be provided.")
        
        if not on_dml_processed:
            def on_process(dml, passed, count):
                print(
                    '.' if passed else 'F', sep='', 
                    end='\n' if count % 100 == 0 else ''
                )
            on_dml_processed = on_process
        
        # storage for operation results summary
        results = _(failed=0, passed=0, errors=[])
        count, cursor = (0, conn.corsor())
        for dml in dml_provider():
            try:
                cursor.execute(dml)
                results.passed += 1
                passed = True
            except Exception as ex:
                results.errors.append([ex, dml])
                results.failed += 1
                passed = False
            
            count += 1
            if count % commit_interval == 0:
                conn.commit()
            
            on_dml_processed(dml, passed, count)
        
        # commit orphaned transactions
        conn.commit()

        # return 
        return results


class XlSheet:
    """
    Represents a light wrapper around openpyxl's Worksheet object. Provides
    convenient ways of iterating over rows which are presented as tuples.
    """
    max_row_check = 10
    
    def __init__(self, source, sheet_name, row_offset=0, col_offset=0):
        # burying import here scopes dependency on openpyxl to just XlSheet
        # this module as a whole doesn't have to depend on openpyxl...
        import openpyxl
        
        workbook = source if type(source) is openpyxl.Workbook else None
        if not workbook and type(source) is str:
            if not os.path.isfile(source):
                raise OSError(source)
            workbook = openpyxl.load_workbook(source, read_only=False)
        
        if not workbook:
            raise ValueError(
                "Expected types for source: str, Workbook. Type provided: %s" %
                (type(source),)
            )
        
        workbook_sheet_names = workbook.get_sheet_names()
        if not sheet_name in workbook_sheet_names:
            raise ValueError(
                "Sheet by the name '%s' not found in '%s'." %
                (sheet_name, workbook_sheet_names)
            )
        
        self.worksheet = workbook[sheet_name]
        self.sheet_name = sheet_name
        self.__col_offset = col_offset
        self.__row_offset = row_offset
        self.__generator = None
        self.__current = None
    
    @property
    def current(self):
        return self.__current
    
    @property
    def max_column(self):
        return self.worksheet.max_column
    
    @property
    def max_row(self):
        return self.worksheet.max_row
    
    @property
    def col_offset(self):
        return self.__col_offset
    
    @col_offset.setter
    def col_offset(self, value):
        self.__col_offset = value
        self.__generator = None
    
    @property
    def row_offset(self):
        return self.__row_offset
    
    @row_offset.setter
    def row_offset(self, value):
        self.__row_offset = value
        self.__generator = None
    
    def iter_rows(self, row_offset=0):
        self.row_offset = row_offset
        return self.__get_generator()
    
    def next(self):
        return self.__get_generator().send(None)
    
    def reset(self):
        self.__generator = None
    
    def __iter__(self):
        return self.__get_generator()
    
    def __get_generator(self):
        def make_generator():
            for i in range(self.row_offset + 1, self.max_row + 1):
                row = []
                for j in range(self.col_offset + 1, self.max_column + 1):
                    value = self.worksheet.cell(row=i, column=j).value
                    row.append(value)
                
                self.__current = row = tuple(row)
                yield row
                
        if self.__generator is None:
            self.__generator = make_generator()
        return self.__generator
    
    @staticmethod
    def find_headers(xlsheet, sample_headers, row_offset=0):
        norm_hdrs = [h.lower() for h in sample_headers]
        hdr_count = len(norm_hdrs)
        
        idx = row_offset
        xlsheet.row_offset = row_offset
        for row in xlsheet:
            idx += 1
            
            norm_row = [str(c or '').lower() for c in row]
            if norm_row[:hdr_count] == norm_hdrs:
                return idx
            
            if (idx - row_offset) == XlSheet.max_row_check:
                return -1


class TypedXlReaderBase(metaclass=ABCMeta):
    
    def __init__(self, xlsheet):
        if not xlsheet:
            raise ValueError("xlsheet cannot be null")
        self.xlsheet = xlsheet

    @abstractmethod
    def _get_column_name_index_map(self):
        pass
        
    @abstractmethod
    def get_rows(self):
        pass

    def _ensure_headers_exists(self, sample_headers, row_offset=0):
        """
        Returns the index at which provided headers are found, otherwise an
        exception is raised.
        """
        index = XlSheet.find_headers(self.xlsheet, sample_headers, row_offset)
        if index == -1:
            raise Exception("Headers not found: %s" % str(sample_headers))
        return index
    
    def _get_data(self, row, key, default=''):
        return row[key] or default
